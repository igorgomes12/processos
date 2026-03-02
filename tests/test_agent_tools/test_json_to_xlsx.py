"""
Testes unitários para agent_tools/json_to_xlsx.py.
Cobre: _stringify_cell, json_to_xlsx.
"""
from __future__ import annotations

import json
import os
from pathlib import Path

import pytest
import pandas as pd
from openpyxl import load_workbook

from agent_tools.json_to_xlsx import _stringify_cell, json_to_xlsx


# ─── _stringify_cell ─────────────────────────────────────────────────────────

class TestStringifyCell:
    def test_none_returns_empty_string(self):
        assert _stringify_cell(None) == ""

    def test_empty_list_returns_empty_array_as(self):
        assert _stringify_cell([], empty_array_as="") == ""
        assert _stringify_cell([], empty_array_as="N/A") == "N/A"

    def test_list_with_one_item(self):
        assert _stringify_cell(["item1"]) == "item1"

    def test_list_with_multiple_items_joined_by_newline(self):
        result = _stringify_cell(["a", "b", "c"])
        assert result == "a\nb\nc"

    def test_list_with_custom_joiner(self):
        result = _stringify_cell(["a", "b"], joiner=", ")
        assert result == "a, b"

    def test_string_returned_as_is(self):
        assert _stringify_cell("hello") == "hello"

    def test_integer_returned_as_is(self):
        assert _stringify_cell(42) == 42

    def test_boolean_returned_as_is(self):
        assert _stringify_cell(True) is True

    def test_list_skips_none_items(self):
        result = _stringify_cell(["a", None, "b"])
        assert result == "a\nb"

    def test_list_converts_non_strings_to_str(self):
        result = _stringify_cell([1, 2, 3])
        assert result == "1\n2\n3"

    def test_nested_list_with_zeros_does_not_skip(self):
        # zero é falsy mas não é None — deve ser incluído
        result = _stringify_cell([0, 1, 2])
        assert "0" in result


# ─── json_to_xlsx ─────────────────────────────────────────────────────────────

def _write_json(tmp_path, data: dict) -> str:
    p = tmp_path / "input.json"
    p.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
    return str(p)


def _xlsx_out(tmp_path) -> str:
    return str(tmp_path / "output.xlsx")


class TestJsonToXlsx:
    def test_creates_xlsx_file(self, tmp_path, json_file, xlsx_path):
        json_to_xlsx(json_file, xlsx_path)
        assert os.path.exists(xlsx_path)

    def test_xlsx_has_correct_sheet_name(self, tmp_path, json_file, xlsx_path):
        json_to_xlsx(json_file, xlsx_path)
        wb = load_workbook(xlsx_path)
        assert "Processos" in wb.sheetnames

    def test_header_row_contains_labels(self, tmp_path, json_file, xlsx_path):
        json_to_xlsx(json_file, xlsx_path)
        wb = load_workbook(xlsx_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert "Frente (N0)" in headers
        assert "Macro processo (N1)" in headers

    def test_data_row_contains_values(self, tmp_path, json_file, xlsx_path):
        json_to_xlsx(json_file, xlsx_path)
        wb = load_workbook(xlsx_path)
        ws = wb.active
        row2 = [cell.value for cell in ws[2]]
        assert "Frente A" in row2
        assert "Macro 1" in row2

    def test_array_field_joined_by_newline(self, tmp_path, json_file, xlsx_path):
        json_to_xlsx(json_file, xlsx_path)
        wb = load_workbook(xlsx_path)
        ws = wb.active
        # Coluna "entradas" deve ter "Entrada 1\nEntrada 2"
        headers = [cell.value for cell in ws[1]]
        entradas_col = headers.index("Entradas") + 1
        cell_value = ws.cell(row=2, column=entradas_col).value
        assert cell_value == "Entrada 1\nEntrada 2"

    def test_empty_array_field_is_empty_string(self, tmp_path):
        data = {
            "columns": [
                {"key": "N0", "label": "N0"},
                {"key": "entradas", "label": "Entradas"},
            ],
            "rows": [{"N0": "X", "entradas": []}],
            "exportHints": {"arrayJoiner": "\n", "emptyArrayAs": ""},
        }
        jp = _write_json(tmp_path, data)
        xp = _xlsx_out(tmp_path)
        json_to_xlsx(jp, xp)
        wb = load_workbook(xp)
        ws = wb.active
        # entradas vazio deve ficar None ou "" (openpyxl armazena None para células vazias)
        val = ws.cell(row=2, column=2).value
        assert val is None or val == ""

    def test_raises_on_missing_columns(self, tmp_path):
        data = {"columns": [], "rows": [{"N0": "X"}]}
        jp = _write_json(tmp_path, data)
        xp = _xlsx_out(tmp_path)
        with pytest.raises(ValueError, match="columns"):
            json_to_xlsx(jp, xp)

    def test_raises_on_none_rows(self, tmp_path):
        data = {
            "columns": [{"key": "N0", "label": "N0"}],
            "rows": None,
        }
        jp = _write_json(tmp_path, data)
        xp = _xlsx_out(tmp_path)
        with pytest.raises((ValueError, TypeError)):
            json_to_xlsx(jp, xp)

    def test_custom_sheet_name(self, tmp_path):
        data = {
            "meta": {"output": {"recommendedSheetName": "MeuProcesso"}},
            "columns": [{"key": "N0", "label": "N0"}],
            "rows": [{"N0": "X"}],
            "exportHints": {},
        }
        jp = _write_json(tmp_path, data)
        xp = _xlsx_out(tmp_path)
        json_to_xlsx(jp, xp)
        wb = load_workbook(xp)
        assert "MeuProcesso" in wb.sheetnames

    def test_freeze_pane_set(self, tmp_path, json_file, xlsx_path):
        json_to_xlsx(json_file, xlsx_path)
        wb = load_workbook(xlsx_path)
        ws = wb.active
        assert ws.freeze_panes == "A2"

    def test_autofilter_set(self, tmp_path, json_file, xlsx_path):
        json_to_xlsx(json_file, xlsx_path)
        wb = load_workbook(xlsx_path)
        ws = wb.active
        assert ws.auto_filter.ref is not None

    def test_header_font_is_bold_white(self, tmp_path, json_file, xlsx_path):
        json_to_xlsx(json_file, xlsx_path)
        wb = load_workbook(xlsx_path)
        ws = wb.active
        header_cell = ws["A1"]
        assert header_cell.font.bold is True
        # openpyxl retorna cor como ARGB (8 hex chars); os últimos 6 chars são o RGB
        assert header_cell.font.color.rgb[-6:].upper() == "FFFFFF"

    def test_multiple_rows(self, tmp_path):
        data = {
            "columns": [{"key": "N0", "label": "N0"}, {"key": "N1", "label": "N1"}],
            "rows": [
                {"N0": "A", "N1": "1"},
                {"N0": "B", "N1": "2"},
                {"N0": "C", "N1": "3"},
            ],
            "exportHints": {},
        }
        jp = _write_json(tmp_path, data)
        xp = _xlsx_out(tmp_path)
        json_to_xlsx(jp, xp)
        wb = load_workbook(xp)
        ws = wb.active
        assert ws.max_row == 4  # 1 cabeçalho + 3 dados

    def test_missing_key_in_row_defaults_to_empty(self, tmp_path):
        data = {
            "columns": [{"key": "N0", "label": "N0"}, {"key": "N1", "label": "N1"}],
            "rows": [{"N0": "A"}],  # N1 ausente
            "exportHints": {},
        }
        jp = _write_json(tmp_path, data)
        xp = _xlsx_out(tmp_path)
        json_to_xlsx(jp, xp)
        wb = load_workbook(xp)
        ws = wb.active
        val = ws.cell(row=2, column=2).value
        assert val is None or val == ""
