import json
import os
from typing import Any, Dict, List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _stringify_cell(value: Any, joiner: str = "\n", empty_array_as: str = "") -> Any:
    """
    Converte valores do JSON para um valor apropriado de célula.
    - Listas viram string juntando pelo joiner
    - None vira string vazia
    """
    if value is None:
        return ""
    if isinstance(value, list):
        if len(value) == 0:
            return empty_array_as
        # garante que todos os itens sejam strings
        return joiner.join(str(x) for x in value if x is not None)
    return value


def json_to_xlsx(json_path: str, xlsx_path: str) -> None:
    # --- Carregar JSON ---
    with open(json_path, "r", encoding="utf-8") as f:
        doc = json.load(f)

    columns: List[Dict[str, str]] = doc.get("columns", [])
    rows: List[Dict[str, Any]] = doc.get("rows", [])
    export_hints: Dict[str, Any] = doc.get("exportHints", {})

    joiner = export_hints.get("arrayJoiner", "\n")
    empty_array_as = export_hints.get("emptyArrayAs", "")

    # Nome da aba (se existir)
    sheet_name = (
        doc.get("meta", {})
           .get("output", {})
           .get("recommendedSheetName", "Processos")
    )

    if not columns:
        raise ValueError("JSON sem 'columns'. Não sei quais colunas escrever no Excel.")
    if rows is None:
        raise ValueError("JSON sem 'rows'.")

    # Ordem das chaves e labels
    col_keys = [c["key"] for c in columns]
    col_labels = [c.get("label", c["key"]) for c in columns]

    # --- Construir DataFrame respeitando ordem das colunas ---
    normalized = []
    for r in rows:
        record = {}
        for k in col_keys:
            record[k] = _stringify_cell(r.get(k), joiner=joiner, empty_array_as=empty_array_as)
        normalized.append(record)

    df = pd.DataFrame(normalized, columns=col_keys)

    # --- Escrever Excel (1ª linha = header com labels) ---
    # Criamos um DataFrame auxiliar só para escrever o header como labels
    # mantendo o conteúdo por keys.
    df_to_write = df.copy()
    df_to_write.columns = col_labels

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df_to_write.to_excel(writer, index=False, sheet_name=sheet_name)

    # --- Pós-formatação com openpyxl ---
    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name]

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")  # azul escuro
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    header_alignment = Alignment(wrap_text=True, vertical="center")

    # Aplica estilo no header
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Wrap text no corpo (útil para campos longos e arrays joinados por \n)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = wrap_alignment

    # Freeze pane (trava header)
    ws.freeze_panes = "A2"

    # Autofilter
    ws.auto_filter.ref = ws.dimensions

    # Ajuste simples de largura (heurístico)
    # limita a largura para evitar colunas absurdas
    MAX_WIDTH = 60
    MIN_WIDTH = 12

    for col_idx, col_name in enumerate(col_labels, start=1):
        letter = get_column_letter(col_idx)
        # calcula maior tamanho entre header e algumas linhas (amostra) para performance
        max_len = len(str(col_name)) if col_name else 0
        sample_limit = min(ws.max_row, 200)  # amostra até 200 linhas
        for r in range(2, sample_limit + 1):
            v = ws.cell(row=r, column=col_idx).value
            if v is None:
                continue
            v_str = str(v)
            # considera apenas a maior linha do texto (quando tem \n)
            v_str = max(v_str.split("\n"), key=len) if "\n" in v_str else v_str
            max_len = max(max_len, len(v_str))

        width = max(MIN_WIDTH, min(MAX_WIDTH, max_len + 2))
        ws.column_dimensions[letter].width = width

    wb.save(xlsx_path)
    wb.close()


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Gera XLSX a partir do JSON canônico (rows/columns).")
    parser.add_argument("json_path", help="Caminho do arquivo .json (ex.: processos_n0_n4_v2.json)")
    parser.add_argument("xlsx_path", nargs="?", help="Saída .xlsx (ex.: processos.xlsx). Opcional.")
    args = parser.parse_args()

    json_path = args.json_path
    xlsx_path = args.xlsx_path or os.path.splitext(json_path)[0] + ".xlsx"

    json_to_xlsx(json_path, xlsx_path)
    print(f"OK: Planilha gerada em: {xlsx_path}")